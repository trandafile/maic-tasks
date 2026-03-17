import os
import random
import re
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import Client, create_client


EXCEL_FILE = "MAIC LAB task manager.xlsx"

# ── Status/Priority mappings ───────────────────────────────────────────────────

VALID_STATUSES = {"Not started", "Working on", "Blocked", "Completed", "Cancelled"}

STATUS_MAP = {
    "planned":      "Not started",
    "notstarted":   "Not started",
    "todo":         "Not started",
    "da fare":      "Not started",
    "workingon":    "Working on",
    "in corso":     "Working on",
    "inprogress":   "Working on",
    "inprogresso":  "Working on",
    "wip":          "Working on",
    "blocked":      "Blocked",
    "bloccato":     "Blocked",
    "completed":    "Completed",
    "completato":   "Completed",
    "done":         "Completed",
    "cancelled":    "Cancelled",
    "canceled":     "Cancelled",
    "annullato":    "Cancelled",
}

VALID_PRIORITIES = {"none", "low", "medium", "high", "urgent"}

PRIORITY_MAP = {
    "nessuna":  "none",
    "bassa":    "low",
    "media":    "medium",
    "alta":     "high",
    "urgente":  "urgent",
}


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text).strip().lower())


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _norm_status(raw) -> str | None:
    """Normalise an Excel status value to one of the 5 valid app statuses."""
    if not raw:
        return None
    key = _norm(str(raw))
    if raw in VALID_STATUSES:
        return raw
    return STATUS_MAP.get(key, "Not started")


def _norm_priority(raw) -> str:
    """Normalise an Excel priority value to one of the 5 valid app priorities."""
    if not raw:
        return "none"
    key = _norm(str(raw))
    if raw in VALID_PRIORITIES:
        return raw
    return PRIORITY_MAP.get(key, "none")


def _read_streamlit_secrets():
    """Small parser for .streamlit/secrets.toml without extra dependencies."""
    secrets_path = Path(".streamlit") / "secrets.toml"
    data = {}
    if not secrets_path.exists():
        return data
    for line in secrets_path.read_text(encoding="utf-8").splitlines():
        raw = line.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        data[key] = value
    return data


def _get_supabase_client() -> Client:
    load_dotenv()
    secrets = _read_streamlit_secrets()
    url = os.environ.get("SUPABASE_URL") or secrets.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY") or secrets.get("SUPABASE_KEY")
    if not url or not key:
        raise ValueError(
            "SUPABASE_URL e SUPABASE_KEY non trovate. Impostale in .env o .streamlit/secrets.toml"
        )
    return create_client(url, key)


def _sheet_rows(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def _pick(row: dict, aliases):
    """Return the first non-None cleaned value for any of the given column aliases.

    FIX: the original returned immediately on the first matching key, even if
    the value was None. Now it continues trying the next alias.
    """
    normalized = {_norm(k): v for k, v in row.items()}
    for alias in aliases:
        key = _norm(alias)
        if key in normalized:
            val = _clean(normalized[key])
            if val is not None:
                return val
    return None


def _avatar_color() -> str:
    return "#{:06x}".format(random.randint(0, 0xFFFFFF))


# ── Clear ──────────────────────────────────────────────────────────────────────

def clear_database(supabase: Client):
    print("[INFO] Cancellazione dati esistenti (ordine child -> parent)...")
    delete_plan = [
        ("comments",          "id"),
        ("task_labels",       "task_id"),
        ("task_dependencies", "task_id"),
        ("subtasks",          "id"),
        ("tasks",             "id"),
        ("deliverables",      "id"),
        ("projects",          "id"),
        ("users",             "email"),
    ]
    for table, pk in delete_plan:
        try:
            rows = supabase.table(table).select(pk).execute().data or []
            if not rows:
                print(f"  [OK] {table}: nessun record")
                continue
            values = [r[pk] for r in rows if r.get(pk) is not None]
            for i in range(0, len(values), 200):
                batch = values[i: i + 200]
                supabase.table(table).delete().in_(pk, batch).execute()
            print(f"  [OK] {table}: eliminati {len(values)} record")
        except Exception as exc:
            print(f"  [WARN] {table}: {exc}")


# ── Import ─────────────────────────────────────────────────────────────────────

def import_data(excel_file: str = EXCEL_FILE):
    supabase = _get_supabase_client()
    wb = load_workbook(filename=excel_file, data_only=True)

    user_map:    dict[str, str] = {}   # name → email, email → email
    project_map: dict[str, int] = {}   # name/acronym → project_id
    identifier_map: dict[int, str] = {}  # project_id → identifier (for sequence_id)
    task_map:    dict[tuple, int] = {}  # (project_name, task_name) → task_id

    # ── Users ────────────────────────────────────────────────────────────────
    print("[INFO] Import utenti da sheet 'Data'...")
    for row in _sheet_rows(wb, "Data"):
        name  = _pick(row, ["People", "Name", "Nome"])
        email = _pick(row, ["E-mail addresses", "Email", "E-mail", "Mail"])
        if not name:
            continue
        if not email:
            email = f"{name.replace(' ', '.').lower()}@example.com"
        payload = {
            "email":        email,
            "name":         name,
            "role":         "user",
            "is_approved":  True,
            "avatar_color": _avatar_color(),
        }
        try:
            supabase.table("users").upsert(payload, on_conflict="email").execute()
            user_map[name]  = email
            user_map[email] = email
            print(f"  [OK] utente: {name} <{email}>")
        except Exception as exc:
            print(f"  [WARN] utente '{name}': {exc}")

    # ── Projects ─────────────────────────────────────────────────────────────
    print("[INFO] Import progetti da sheet 'Projects'...")
    for row in _sheet_rows(wb, "Projects"):
        name    = _pick(row, ["Nome progetto", "Project name", "Project", "Name"])
        acronym = _pick(row, ["Acronimo", "Acronym"])
        if not name:
            name = acronym
        if not name:
            continue

        # identifier = acronym uppercased (used as prefix for sequence_id)
        identifier = (acronym or name).upper().replace(" ", "")

        payload = {
            "name":           name,
            "acronym":        acronym,
            "identifier":     identifier,
            "funding_agency": _pick(row, ["Ente finanziatore", "Funding agency", "Funder"]),
            "start_date":     _parse_date(_pick(row, ["Start date", "Start", "Data inizio"])),
            "end_date":       _parse_date(_pick(row, ["End date", "End", "Data fine"])),
            "is_archived":    False,
        }
        try:
            res = supabase.table("projects").insert(payload).execute()
            if not res.data:
                continue
            project_id = res.data[0]["id"]
            project_map[name]       = project_id
            identifier_map[project_id] = identifier
            if acronym:
                project_map[acronym] = project_id
                project_map[acronym.strip()] = project_id
            print(f"  [OK] progetto: {name} ({identifier})")
        except Exception as exc:
            print(f"  [WARN] progetto '{name}': {exc}")

    # ── Tasks ─────────────────────────────────────────────────────────────────
    def import_tasks_from_sheet(sheet_name: str, archived: bool):
        print(f"[INFO] Import task da sheet '{sheet_name}'...")
        sort_counter: dict[int, int] = {}

        for row in _sheet_rows(wb, sheet_name):
            project_name = _pick(row, ["Project", "Progetto"])
            task_name    = _pick(row, ["Task name", "Task", "Nome task"])
            if not project_name or not task_name:
                continue

            # Fuzzy project lookup: try exact match first, then stripped
            project_id = project_map.get(project_name) or project_map.get(project_name.strip())
            if not project_id:
                print(f"  [WARN] task '{task_name}': progetto '{project_name}' non trovato")
                continue

            sort_counter[project_id] = sort_counter.get(project_id, 0) + 10

            owner_key      = _pick(row, ["Task owner", "Owner"])
            supervisor_key = _pick(row, ["Task supervisor", "Supervisor"])

            # Notes: try short notes first, fall back to extended notes link
            notes = _pick(row, ["Short notes", "Notes", "Extended notes (link)"])

            payload = {
                "project_id":       project_id,
                "name":             task_name,
                "owner_email":      user_map.get(owner_key) if owner_key else None,
                "supervisor_email": user_map.get(supervisor_key) if supervisor_key else None,
                "status":           _norm_status(_pick(row, ["Task status", "Status"])),
                "priority":         _norm_priority(_pick(row, ["Priority", "Priorita", "Priorità"])),
                "deadline":         _parse_date(_pick(row, ["Expected deadline", "Deadline"])),
                "completion_date":  _parse_date(_pick(row, ["Completion date"])),
                "notes":            notes,
                "sort_order":       sort_counter[project_id],
                "is_archived":      archived,
            }

            try:
                res = supabase.table("tasks").insert(payload).execute()
                if not res.data:
                    continue
                task_id = res.data[0]["id"]
                task_map[(project_name, task_name)]        = task_id
                task_map[(project_name.strip(), task_name)] = task_id

                # Generate sequence_id: {IDENTIFIER}-{task_id}
                ident = identifier_map.get(project_id, "TSK")
                seq_id = f"{ident}-{task_id}"
                supabase.table("tasks").update({"sequence_id": seq_id}).eq("id", task_id).execute()

                print(f"  [OK] task: {seq_id} — {task_name}")
            except Exception as exc:
                print(f"  [WARN] task '{task_name}': {exc}")

    import_tasks_from_sheet("Tasks",         archived=False)
    import_tasks_from_sheet("Archivio Tasks", archived=True)

    # ── Subtasks ──────────────────────────────────────────────────────────────
    def import_subtasks_from_sheet(sheet_name: str, archived: bool):
        print(f"[INFO] Import subtask da sheet '{sheet_name}'...")
        sort_counter: dict[int, int] = {}

        for row in _sheet_rows(wb, sheet_name):
            project_name  = _pick(row, ["Project", "Progetto"])
            task_name     = _pick(row, ["Task name", "Task"])
            subtask_name  = _pick(row, ["Sub-task name", "Subtask", "Sub task", "Nome subtask"])
            if not project_name or not task_name or not subtask_name:
                continue

            # Fuzzy task lookup
            parent_id = (
                task_map.get((project_name, task_name))
                or task_map.get((project_name.strip(), task_name))
            )
            if not parent_id:
                print(
                    f"  [WARN] subtask '{subtask_name}': "
                    f"task parent '{task_name}' non trovato in '{project_name}'"
                )
                continue

            sort_counter[parent_id] = sort_counter.get(parent_id, 0) + 10

            owner_key      = _pick(row, ["Sub-Task owner", "Subtask owner", "Owner"])
            supervisor_key = _pick(row, ["Subtask supervisor", "Sub-Task supervisor", "Supervisor"])

            notes = _pick(row, ["Short notes", "Notes", "Extended notes (link)"])

            payload = {
                "task_id":          parent_id,
                "name":             subtask_name,
                "owner_email":      user_map.get(owner_key) if owner_key else None,
                "supervisor_email": user_map.get(supervisor_key) if supervisor_key else None,
                "status":           _norm_status(_pick(row, ["Sub-Task status", "Status"])),
                "deadline":         _parse_date(_pick(row, ["Deadline", "Expected deadline"])),
                "notes":            notes,
                "sort_order":       sort_counter[parent_id],
                "is_archived":      archived,
            }

            try:
                supabase.table("subtasks").insert(payload).execute()
                print(f"  [OK] subtask: {subtask_name} (parent task #{parent_id})")
            except Exception as exc:
                print(f"  [WARN] subtask '{subtask_name}': {exc}")

    import_subtasks_from_sheet("SubTasks",          archived=False)
    import_subtasks_from_sheet("Archivio SubTasks",  archived=True)

    print("\n[OK] Import completato.")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    workbook_path = Path(EXCEL_FILE)
    if not workbook_path.exists():
        print(f"[ERRORE] File Excel non trovato: {workbook_path.resolve()}")
        raise SystemExit(1)

    conferma = input(
        "\nQuesta operazione CANCELLERA' tutti i dati da Supabase e importera' da Excel.\n"
        "Digita CONFERMA per procedere: "
    )
    if conferma != "CONFERMA":
        print("Operazione annullata.")
        raise SystemExit(0)

    client = _get_supabase_client()
    clear_database(client)
    import_data(EXCEL_FILE)
