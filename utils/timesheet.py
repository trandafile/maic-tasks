"""utils/timesheet.py — Monthly timesheet: autofill + Excel export.

The Excel layout mirrors the MIUR template already in use
(``periodo_YYYY-MM_prog<CUP>``):

    A3  TIMESHEET PER RENDICONTAZIONE PERSONALE:
    A4  Anno: | I4 <year>            | Q4 Mese: | Y4 <italian month>
    A5  Cognome: | I5 <surname>      | Q5 Nome: | Y5 <first name>
    A6  Codice Fiscale: | I6 <cf>
    A7  CUP del progetto: | B7 <cup>
    A8  Soggetto attuatore: ...
    A9  Titolo del progetto: ...
    A10 Tipo del progetto: ...
    A11 Monte ore lavorative annuo previsto: ...
    A12 header row: "Attività svolta sul Progetto\\Day" + days 1..N
    A13..  one row per configurable project activity (+ " - <CUP>" when the
           row is imputable to the project), then a "Totale" row
    then    Firma Incaricato / Data / Firma

Totals are written as literal values (not formulas) so that tools reading the
file with pandas/openpyxl in data-only mode see numbers, exactly like the
existing "TIMESHEET GENERATO DA REPORT" files.
"""

from __future__ import annotations

import calendar
import datetime
import math
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MONTHS_IT = {
    1: "Gennaio", 2: "Febbraio", 3: "Marzo", 4: "Aprile",
    5: "Maggio", 6: "Giugno", 7: "Luglio", 8: "Agosto",
    9: "Settembre", 10: "Ottobre", 11: "Novembre", 12: "Dicembre",
}

# Hours are allocated in half-hour units so an odd split of e.g. 8h over three
# activities still sums back exactly to 8h.
_UNITS_PER_HOUR = 2


def days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def is_working_day(d: datetime.date) -> bool:
    """Monday–Friday. Public holidays are not modelled (edit the grid by hand)."""
    return d.weekday() < 5


def working_days(year: int, month: int) -> list[int]:
    return [
        d for d in range(1, days_in_month(year, month) + 1)
        if is_working_day(datetime.date(year, month, d))
    ]


def _parse_date(value) -> datetime.date | None:
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _tidy(h: float) -> float | int:
    """Return an int when the value is integral (keeps the sheet clean)."""
    return int(h) if float(h).is_integer() else round(float(h), 2)


def split_daily_hours(daily_hours: float, shares: list[float]) -> list[float]:
    """Split `daily_hours` across `shares` (percentages) using largest remainder.

    Guarantees ``sum(result) == daily_hours`` whenever daily_hours is a multiple
    of 0.5. Rows with a zero share never receive hours.
    """
    total_share = sum(s for s in shares if s > 0)
    if total_share <= 0 or daily_hours <= 0:
        return [0.0] * len(shares)

    units_total = int(round(float(daily_hours) * _UNITS_PER_HOUR))
    raw = [(units_total * s / total_share) if s > 0 else 0.0 for s in shares]
    base = [int(math.floor(x)) for x in raw]
    remainder = units_total - sum(base)

    # Distribute the leftover units to the biggest fractional parts, but only
    # among rows that actually carry a share.
    eligible = [i for i, s in enumerate(shares) if s > 0]
    eligible.sort(key=lambda i: (raw[i] - base[i], shares[i]), reverse=True)
    for k in range(remainder):
        base[eligible[k % len(eligible)]] += 1

    return [b / _UNITS_PER_HOUR for b in base]


def autofill_grid(
    activities: list[dict],
    daily_hours: float,
    year: int,
    month: int,
    start_date=None,
    end_date=None,
) -> dict[str, dict[str, float]]:
    """Book `daily_hours` on every working day inside the contract period.

    Returns {activity_id: {day: hours}} with string keys (JSONB friendly).
    Days outside the contract period, weekends, and zero-share rows stay empty.
    """
    start = _parse_date(start_date)
    end = _parse_date(end_date)
    shares = [float(a.get("default_share_pct") or 0) for a in activities]

    grid: dict[str, dict[str, float]] = {str(a["id"]): {} for a in activities}
    if sum(shares) <= 0:
        return grid

    for day in range(1, days_in_month(year, month) + 1):
        d = datetime.date(year, month, day)
        if not is_working_day(d):
            continue
        if start and d < start:
            continue
        if end and d > end:
            continue
        for a, hours in zip(activities, split_daily_hours(daily_hours, shares)):
            if hours > 0:
                grid[str(a["id"])][str(day)] = _tidy(hours)
    return grid


def grid_cell(grid: dict, activity_id, day: int) -> float:
    try:
        return float(grid.get(str(activity_id), {}).get(str(day), 0) or 0)
    except (TypeError, ValueError):
        return 0.0


def row_total(grid: dict, activity_id, year: int, month: int) -> float:
    return sum(grid_cell(grid, activity_id, d) for d in range(1, days_in_month(year, month) + 1))


def day_total(grid: dict, activities: list[dict], day: int) -> float:
    return sum(grid_cell(grid, a["id"], day) for a in activities)


def month_total(grid: dict, activities: list[dict], year: int, month: int) -> float:
    return sum(row_total(grid, a["id"], year, month) for a in activities)


def activity_label(activity: dict, cup: str | None) -> str:
    """Rows imputable to the project carry the CUP, as in the MIUR template."""
    name = (activity.get("name") or "").strip()
    if activity.get("counts_to_project") and cup:
        return f"{name} - {cup}"
    return name


def split_person_name(full_name: str) -> tuple[str, str]:
    """(surname, first name) from "Emilio Arnieri" → ("ARNIERI", "EMILIO").

    Heuristic: the first token is the given name, the rest is the surname —
    which matches how names are stored in the users table.
    """
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0].upper(), ""
    return " ".join(parts[1:]).upper(), parts[0].upper()


# ── Excel ─────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL = PatternFill("solid", fgColor="DDE3EA")
_TOT_FILL = PatternFill("solid", fgColor="F1F3F5")
_WEEKEND_FILL = PatternFill("solid", fgColor="F6F6F6")


def sheet_title(year: int, month: int, cup: str | None) -> str:
    """Excel sheet names cap at 31 chars — the same truncation the originals show."""
    return f"periodo_{year}-{month:02d}_prog{(cup or '').strip()}"[:31]


def build_timesheet_excel(
    *,
    user: dict,
    contract: dict,
    project: dict,
    activities: list[dict],
    grid: dict,
    year: int,
    month: int,
) -> BytesIO:
    n_days = days_in_month(year, month)
    last_col = 2 + n_days                      # A + days → total column
    cup = (project.get("cup") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title(year, month, cup)

    surname, firstname = split_person_name(user.get("name", ""))
    bold = Font(bold=True)

    def put(coord: str, value, *, bold_it: bool = False):
        c = ws[coord]
        c.value = value
        if bold_it:
            c.font = bold
        return c

    def merge(row: int, c1: int, c2: int):
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

    # ── Header ────────────────────────────────────────────────────────────────
    put("A3", "TIMESHEET PER RENDICONTAZIONE PERSONALE:", bold_it=True)
    merge(3, 1, last_col)

    put("A4", "Anno:", bold_it=True); put("I4", year)
    put("Q4", "Mese:", bold_it=True); put("Y4", MONTHS_IT[month])
    merge(4, 1, 8); merge(4, 9, 16); merge(4, 17, 24); merge(4, 25, last_col)

    put("A5", "Cognome:", bold_it=True); put("I5", surname)
    put("Q5", "Nome:", bold_it=True); put("Y5", firstname)
    merge(5, 1, 8); merge(5, 9, 16); merge(5, 17, 24); merge(5, 25, last_col)

    put("A6", "Codice Fiscale:", bold_it=True); put("I6", user.get("fiscal_code") or "")
    merge(6, 1, 8); merge(6, 9, last_col)

    rows_meta = [
        ("CUP del progetto:", cup),
        ("Soggetto attuatore:", project.get("soggetto_attuatore") or ""),
        ("Titolo del progetto:", project.get("name") or ""),
        ("Tipo del progetto:", project.get("project_type") or ""),
        ("Monte ore lavorative annuo previsto:", contract.get("annual_hours") or ""),
    ]
    for i, (label, value) in enumerate(rows_meta, start=7):
        put(f"A{i}", label, bold_it=True)
        ws.cell(i, 2).value = value
        merge(i, 2, last_col)

    # ── Grid header ───────────────────────────────────────────────────────────
    hdr_row = 12
    hc = put(f"A{hdr_row}", "Attività svolta sul Progetto\\Day", bold_it=True)
    hc.fill = _HDR_FILL
    hc.border = _BORDER
    for d in range(1, n_days + 1):
        c = ws.cell(hdr_row, 1 + d)
        c.value = d
        c.font = bold
        c.fill = _HDR_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    # The template leaves the total-column header blank; keep it that way.
    ws.cell(hdr_row, last_col).fill = _HDR_FILL
    ws.cell(hdr_row, last_col).border = _BORDER

    weekend_cols = {
        1 + d for d in range(1, n_days + 1)
        if not is_working_day(datetime.date(year, month, d))
    }

    # ── Activity rows ─────────────────────────────────────────────────────────
    r = hdr_row + 1
    for a in activities:
        lc = ws.cell(r, 1)
        lc.value = activity_label(a, cup)
        lc.border = _BORDER
        for d in range(1, n_days + 1):
            c = ws.cell(r, 1 + d)
            h = grid_cell(grid, a["id"], d)
            c.value = _tidy(h) if h else None
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center")
            if (1 + d) in weekend_cols:
                c.fill = _WEEKEND_FILL
        tc = ws.cell(r, last_col)
        tc.value = _tidy(row_total(grid, a["id"], year, month))
        tc.font = bold
        tc.border = _BORDER
        tc.alignment = Alignment(horizontal="center")
        r += 1

    # ── Totale row ────────────────────────────────────────────────────────────
    tot_row = r
    tl = ws.cell(tot_row, 1)
    tl.value = "Totale"
    tl.font = bold
    tl.fill = _TOT_FILL
    tl.border = _BORDER
    for d in range(1, n_days + 1):
        c = ws.cell(tot_row, 1 + d)
        c.value = _tidy(day_total(grid, activities, d))
        c.font = bold
        c.fill = _TOT_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    gt = ws.cell(tot_row, last_col)
    gt.value = _tidy(month_total(grid, activities, year, month))
    gt.font = bold
    gt.fill = _TOT_FILL
    gt.border = _BORDER
    gt.alignment = Alignment(horizontal="center")

    # ── Signature block ───────────────────────────────────────────────────────
    sig = tot_row + 2
    put(f"A{sig}", "Firma Incaricato (Nome e Cognome):", bold_it=True)
    ws.cell(sig, 2).value = f"{firstname} {surname}".strip()
    put(f"A{sig + 1}", "Data:", bold_it=True)
    put(f"A{sig + 2}", "Firma:", bold_it=True)
    merge(sig + 1, 1, 8)
    merge(sig + 2, 1, 8)
    # Footer sits immediately under "Firma:", as in the original files.
    put(f"A{sig + 3}", "TIMESHEET GENERATO DA MAIC LAB TASK MANAGER")
    merge(sig + 3, 1, last_col)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    for d in range(1, n_days + 1):
        ws.column_dimensions[get_column_letter(1 + d)].width = 4.2
    ws.column_dimensions[get_column_letter(last_col)].width = 7
    ws.freeze_panes = ws.cell(hdr_row + 1, 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def excel_filename(user_name: str, year: int, month: int) -> str:
    surname, _ = split_person_name(user_name)
    surname = "".join(ch for ch in surname.title() if ch.isalnum()) or "Timesheet"
    return f"{surname}{year}_{month:02d}.xlsx"
